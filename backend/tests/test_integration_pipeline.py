from __future__ import annotations

import json
import time
from collections.abc import Callable
from io import BytesIO
from uuid import uuid4

import pytest
from confluent_kafka import Consumer
from flask.testing import FlaskClient
from openpyxl import load_workbook
from pytest import MonkeyPatch

from app.core import KAFKA_TOPICS, FailureCode, JobStatus, Settings
from app.core.errors import ApiError, AuthenticationError
from app.db import session_scope
from app.db.repositories import (
    get_current_output_artifact,
    get_job,
    get_processing_attempt,
    get_user_by_firebase_uid,
    list_processing_attempts_for_job,
)
from app.services.storage_service import delete_object, object_exists
from app.services.validation_service import ValidationResult
from app.services.worker_service import process_worker_event

TOKEN_CLAIMS: dict[str, dict[str, object]] = {
    "user-token": {
        "uid": "firebase-user",
        "email": "user@example.com",
        "name": "Regular User",
    },
    "admin-token": {
        "uid": "firebase-admin",
        "email": "admin@example.com",
        "name": "Admin User",
    },
    "other-user-token": {
        "uid": "firebase-other",
        "email": "other@example.com",
        "name": "Other User",
    },
}


def build_upload_data(file_name: str, payload: bytes) -> dict[str, tuple[BytesIO, str]]:
    return {"file": (BytesIO(payload), file_name)}


@pytest.fixture(autouse=True)
def patch_auth_verifier(monkeypatch: MonkeyPatch) -> None:
    def fake_verify(_settings: Settings, bearer_token: str) -> dict[str, object]:
        claims = TOKEN_CLAIMS.get(bearer_token)
        if claims is None:
            raise AuthenticationError("Unknown test token")
        return dict(claims)

    monkeypatch.setattr("app.services.auth_service.verify_bearer_token", fake_verify)


def test_upload_persists_source_job_and_submit_event(
    client: FlaskClient,
    settings: Settings,
    auth_headers: Callable[[str], dict[str, str]],
    invoice_pdf_bytes: bytes,
) -> None:
    response = client.post(
        "/api/uploads",
        headers=auth_headers("user-token"),
        data=build_upload_data("invoice.pdf", invoice_pdf_bytes),
    )

    assert response.status_code == 201
    payload = response.get_json()
    assert payload["status"] == JobStatus.QUEUED

    job_id = str(payload["job_id"])
    with session_scope(settings) as session:
        job = get_job(session, job_id)
        assert job is not None
        assert job.job_status == JobStatus.QUEUED
        assert job.current_stage == "event_published"
        assert job.source_file_id is not None

        source_user = get_user_by_firebase_uid(session, "firebase-user")
        assert source_user is not None
        expected_key = f"receiving/{source_user.id}/{job_id}/source.pdf"
        assert object_exists(settings, key=expected_key) is True

    event = _consume_matching_event(settings, KAFKA_TOPICS.submit, job_id)
    assert event["job_id"] == job_id
    assert event["attempt_type"] == "initial"


def test_worker_processing_creates_artifact_and_downloads_excel(
    client: FlaskClient,
    settings: Settings,
    auth_headers: Callable[[str], dict[str, str]],
    invoice_pdf_bytes: bytes,
) -> None:
    upload_response = client.post(
        "/api/uploads",
        headers=auth_headers("user-token"),
        data=build_upload_data("invoice.pdf", invoice_pdf_bytes),
    )
    job_id = str(upload_response.get_json()["job_id"])
    event = _consume_matching_event(settings, KAFKA_TOPICS.submit, job_id)

    with session_scope(settings) as session:
        result = process_worker_event(
            session,
            settings=settings,
            event_payload=event,
            worker_request_id="worker-success",
        )

    assert result.status == JobStatus.COMPLETED

    detail_response = client.get(f"/api/jobs/{job_id}", headers=auth_headers("user-token"))
    detail_payload = detail_response.get_json()
    assert detail_response.status_code == 200
    assert detail_payload["status"] == JobStatus.COMPLETED
    assert detail_payload["download_available"] is True

    download_response = client.get(
        f"/api/jobs/{job_id}/download",
        headers=auth_headers("user-token"),
    )
    assert download_response.status_code == 200
    assert download_response.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(download_response.data))
    assert "Summary" in workbook.sheetnames
    assert "Traceability" in workbook.sheetnames
    assert "Line Items" in workbook.sheetnames

    with session_scope(settings) as session:
        job = get_job(session, job_id)
        assert job is not None
        artifact = get_current_output_artifact(session, job_id=job_id)
        assert artifact is not None
        assert list_processing_attempts_for_job(session, job_id=job_id)[0].status == "completed"


def test_invalid_pdf_upload_is_rejected(
    client: FlaskClient,
    auth_headers: Callable[[str], dict[str, str]],
) -> None:
    response = client.post(
        "/api/uploads",
        headers=auth_headers("user-token"),
        data=build_upload_data("not-a-pdf.txt", b"plain text"),
    )

    assert response.status_code == 400
    assert response.get_json()["error"]["code"] == FailureCode.UPLOAD_INVALID_TYPE


def test_user_failure_message_is_sanitized_when_gemini_fails(
    client: FlaskClient,
    settings: Settings,
    auth_headers: Callable[[str], dict[str, str]],
    invoice_pdf_bytes: bytes,
    monkeypatch: MonkeyPatch,
) -> None:
    def fail_extract(*args: object, **kwargs: object) -> object:
        raise ApiError(
            code=FailureCode.GEMINI_REQUEST_FAILED,
            message="Traceback: internal gemini failure",
        )

    monkeypatch.setattr("app.services.worker_service.extract_document_data", fail_extract)

    upload_response = client.post(
        "/api/uploads",
        headers=auth_headers("user-token"),
        data=build_upload_data("invoice.pdf", invoice_pdf_bytes),
    )
    job_id = str(upload_response.get_json()["job_id"])
    event = _consume_matching_event(settings, KAFKA_TOPICS.submit, job_id)

    with session_scope(settings) as session:
        with pytest.raises(ApiError):
            process_worker_event(
                session,
                settings=settings,
                event_payload=event,
                worker_request_id="worker-gemini-failure",
            )

    user_detail = client.get(f"/api/jobs/{job_id}", headers=auth_headers("user-token"))
    user_payload = user_detail.get_json()
    assert user_detail.status_code == 200
    assert user_payload["status"] == JobStatus.FAILED
    assert "Traceback" not in str(user_payload["failure_message"])

    admin_detail = client.get(f"/api/admin/jobs/{job_id}", headers=auth_headers("admin-token"))
    assert admin_detail.status_code == 200
    assert admin_detail.get_json()["failure_code"] == FailureCode.GEMINI_REQUEST_FAILED


def test_validation_failure_blocks_artifact_creation(
    client: FlaskClient,
    settings: Settings,
    auth_headers: Callable[[str], dict[str, str]],
    report_pdf_bytes: bytes,
    monkeypatch: MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        "app.services.worker_service.validate_normalized_output",
        lambda _payload: ValidationResult(
            valid=False,
            errors=[{"field": "sections", "message": "No sections detected."}],
        ),
    )

    upload_response = client.post(
        "/api/uploads",
        headers=auth_headers("user-token"),
        data=build_upload_data("report.pdf", report_pdf_bytes),
    )
    job_id = str(upload_response.get_json()["job_id"])
    event = _consume_matching_event(settings, KAFKA_TOPICS.submit, job_id)

    with session_scope(settings) as session:
        with pytest.raises(ApiError):
            process_worker_event(
                session,
                settings=settings,
                event_payload=event,
                worker_request_id="worker-validation-failure",
            )
        assert get_current_output_artifact(session, job_id=job_id) is None

    response = client.get(f"/api/jobs/{job_id}", headers=auth_headers("user-token"))
    assert response.status_code == 200
    assert response.get_json()["status"] == JobStatus.FAILED


def test_download_missing_artifact_returns_not_found(
    client: FlaskClient,
    settings: Settings,
    auth_headers: Callable[[str], dict[str, str]],
    invoice_pdf_bytes: bytes,
) -> None:
    upload_response = client.post(
        "/api/uploads",
        headers=auth_headers("user-token"),
        data=build_upload_data("invoice.pdf", invoice_pdf_bytes),
    )
    job_id = str(upload_response.get_json()["job_id"])
    event = _consume_matching_event(settings, KAFKA_TOPICS.submit, job_id)

    with session_scope(settings) as session:
        process_worker_event(
            session,
            settings=settings,
            event_payload=event,
            worker_request_id="worker-missing-artifact",
        )
        artifact = get_current_output_artifact(session, job_id=job_id)
        assert artifact is not None
        file_record = get_job(session, job_id)
        assert file_record is not None

    with session_scope(settings) as session:
        artifact = get_current_output_artifact(session, job_id=job_id)
        assert artifact is not None
        output_job = get_job(session, job_id)
        assert output_job is not None
        output_key = f"processed/{output_job.user_id}/{job_id}/output.xlsx"
    delete_object(settings, key=output_key)

    response = client.get(f"/api/jobs/{job_id}/download", headers=auth_headers("user-token"))
    assert response.status_code == 404
    assert response.get_json()["error"]["code"] == FailureCode.ARTIFACT_NOT_FOUND


def test_admin_retry_creates_new_attempt_and_retry_event(
    client: FlaskClient,
    settings: Settings,
    auth_headers: Callable[[str], dict[str, str]],
    invoice_pdf_bytes: bytes,
    monkeypatch: MonkeyPatch,
) -> None:
    def fail_extract(*args: object, **kwargs: object) -> object:
        raise ApiError(
            code=FailureCode.GEMINI_REQUEST_FAILED,
            message="Extraction provider unavailable.",
        )

    monkeypatch.setattr("app.services.worker_service.extract_document_data", fail_extract)

    upload_response = client.post(
        "/api/uploads",
        headers=auth_headers("user-token"),
        data=build_upload_data("invoice.pdf", invoice_pdf_bytes),
    )
    job_id = str(upload_response.get_json()["job_id"])
    event = _consume_matching_event(settings, KAFKA_TOPICS.submit, job_id)

    with session_scope(settings) as session:
        with pytest.raises(ApiError):
            process_worker_event(
                session,
                settings=settings,
                event_payload=event,
                worker_request_id="worker-retry-source",
            )

    inspect_response = client.get(f"/api/admin/jobs/{job_id}", headers=auth_headers("admin-token"))
    assert inspect_response.status_code == 200

    retry_response = client.post(
        f"/api/admin/jobs/{job_id}/retry",
        headers=auth_headers("admin-token"),
        json={"notes": "Retry after transient Gemini outage."},
    )
    assert retry_response.status_code == 202
    retry_payload = retry_response.get_json()

    with session_scope(settings) as session:
        attempts = list_processing_attempts_for_job(session, job_id=job_id)
        assert len(attempts) == 2
        queued_attempt = get_processing_attempt(
            session,
            str(retry_payload["processing_attempt_id"]),
        )
        assert queued_attempt is not None
        assert queued_attempt.status == JobStatus.QUEUED

    retry_event = _consume_matching_event(settings, KAFKA_TOPICS.retry, job_id)
    assert retry_event["processing_attempt_id"] == retry_payload["processing_attempt_id"]


def _consume_matching_event(
    settings: Settings,
    topic: str,
    job_id: str,
    timeout_seconds: float = 12.0,
) -> dict[str, object]:
    consumer = Consumer(
        {
            "bootstrap.servers": settings.kafka_bootstrap_servers,
            "group.id": f"pdfextract-tests-{uuid4()}",
            "auto.offset.reset": "earliest",
        }
    )
    consumer.subscribe([topic])
    deadline = time.monotonic() + timeout_seconds

    try:
        while time.monotonic() < deadline:
            message = consumer.poll(1.0)
            if message is None:
                continue
            if message.error():
                raise AssertionError(f"Kafka consume failed: {message.error()}")

            payload = json.loads(message.value().decode("utf-8"))
            if payload.get("job_id") == job_id:
                return payload
    finally:
        consumer.close()

    raise AssertionError(f"Kafka event for job {job_id} was not observed on topic {topic}.")
