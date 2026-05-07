from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core import DocumentType


def generate_excel_workbook(
    *,
    job_id: str,
    source_filename: str,
    normalized_json: dict[str, Any],
) -> tuple[bytes, list[str]]:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _write_summary(summary_sheet, normalized_json)

    traceability_sheet = workbook.create_sheet("Traceability")
    traceability_sheet.append(["Job ID", job_id])
    traceability_sheet.append(["Source Filename", source_filename])
    traceability_sheet.append(["Document Type", normalized_json.get("document_type") or "unknown"])

    document_type = str(normalized_json.get("document_type") or "")
    if document_type == DocumentType.INVOICE:
        _write_invoice_sheet(workbook, normalized_json)
    else:
        _write_report_sheets(workbook, normalized_json)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), workbook.sheetnames


def _write_summary(sheet: Worksheet, normalized_json: dict[str, Any]) -> None:
    sheet.append(["Field", "Value"])
    for key, value in normalized_json.items():
        if isinstance(value, list | dict):
            sheet.append([key, str(value)])
        else:
            sheet.append([key, value])


def _write_invoice_sheet(workbook: Workbook, normalized_json: dict[str, Any]) -> None:
    line_items_sheet = workbook.create_sheet("Line Items")
    line_items_sheet.append(["Description", "Quantity", "Unit Price", "Line Total"])
    for item in normalized_json.get("line_items", []):
        if not isinstance(item, dict):
            continue
        line_items_sheet.append(
            [
                item.get("description"),
                item.get("quantity"),
                item.get("unit_price"),
                item.get("line_total"),
            ]
        )


def _write_report_sheets(workbook: Workbook, normalized_json: dict[str, Any]) -> None:
    sections_sheet = workbook.create_sheet("Sections")
    sections_sheet.append(["Heading", "Content"])
    for section in normalized_json.get("sections", []):
        if not isinstance(section, dict):
            continue
        sections_sheet.append([section.get("heading"), section.get("content")])

    if normalized_json.get("table_count"):
        tables_sheet = workbook.create_sheet("Tables")
        tables_sheet.append(["Table Count"])
        tables_sheet.append([normalized_json.get("table_count")])
